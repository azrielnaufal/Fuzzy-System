from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from itertools import islice
import xlwt
from xlwt import Workbook
 
data = load_workbook('restoran.xlsx')
wsData = data.active

crispList = []
fuzzyInput = {
    "pelayanan" : [],
    "makanan" : []
}
l = {
    "vwp" : [0, 0, 10, 20],
    "wp" : [10, 25, 25, 40],
    "lmp" : [30, 40, 40, 50],
    "hmp" : [40, 55, 55, 70],
    "sp" : [60, 70, 70, 80],
    "vsp" : [70, 80, 80, 90],
    "ep" : [80, 90, 100, 100]
}
l2 = {
    "vwp" : [0, 0, 1, 2],
    "wp" : [1, 2.5, 2.5, 4],
    "lmp" : [3, 4, 4, 5],
    "hmp" : [4, 5.5, 5.5, 7],
    "sp" : [6, 7, 7, 8],
    "vsp" : [7, 8, 8, 9],
    "ep" : [8, 9, 10, 10]
}

for i in range(2, 102):
    cr = {}
    d = wsData.cell(row=i, column=1)
    cr["id"] = d.value
    d = wsData.cell(row=i, column=2)
    cr["pelayanan"] = d.value
    d = wsData.cell(row=i, column=3)
    cr["makanan"] = d.value
    crispList.append(cr)

def trapezoid(id, x, jenis):
    batas = {}
    fuzz = {
        "id" : id,
        "crisp" : x,
        "fuz" : {}
    }
    if jenis == "pelayanan":
        for k,v in l.items():
            if v[0]<=x<=v[-1]:
                batas[k] = v
    elif jenis == "makanan":
        for k,v in l2.items():
            if v[0]<=x<=v[-1]:
                batas[k] = v
    for k, v in batas.items():
        a, b, c, d = v[0], v[1], v[2], v[3]
        if a<=x<=b:
          temp = (x-a)/(b-a)
        elif b<=x<=c:
          temp = 1
        elif c<=x<=d:
          temp = (x-d)/(c-d)
        if temp > 0:
          fuzz["fuz"][k] = temp
    return fuzz

def membershipFunctions():
    for i in crispList:
        fuzzyInput["pelayanan"].append(trapezoid(i["id"], i["pelayanan"], "pelayanan"))
        fuzzyInput["makanan"].append(trapezoid(i["id"], i["makanan"], "makanan"))

membershipFunctions()

def MamdaniInference(id):
  fuzzyOutput = {
      "fuzOut" : {}
  }
  for k,v in fuzzyInput["pelayanan"][id]["fuz"].items():
    for k1,v1 in fuzzyInput["makanan"][id]["fuz"].items():
      sem = fuzzy_rules(k, k1)
      derajat_rendah = min(v,v1)
      fuzzyOutput["id"] = id + 1     
      if len(fuzzyOutput["fuzOut"]) == 0:
        fuzzyOutput["fuzOut"][sem] = derajat_rendah
      else:
        if sem in fuzzyOutput["fuzOut"]:
          fuzzyOutput["fuzOut"][sem] = max(derajat_rendah,fuzzyOutput["fuzOut"][sem])
        else:
          fuzzyOutput["fuzOut"][sem] = derajat_rendah
  return fuzzyOutput

def fuzzy_rules(a,b):
  if a == 'vwp' and b == 'vwp':
    return 'vwp'
  elif a == 'vwp' and b == 'wp':
    return 'vwp'
  elif a == 'vwp' and b == 'lmp':
    return 'wp'
  elif a == 'vwp' and b == 'hmp':
    return 'wp'
  elif a == 'vwp' and b == 'sp':
    return 'lmp'
  elif a == 'vwp' and b == 'vsp':
    return 'hmp'
  elif a == 'vwp' and b == 'ep':
    return 'hmp'
  ######################
  if a == 'wp' and b == 'vwp':
    return 'vwp'
  elif a == 'wp' and b == 'wp':
    return 'vwp'
  elif a == 'wp' and b == 'lmp':
    return 'wp'
  elif a == 'wp' and b == 'hmp':
    return 'lmp'
  elif a == 'wp' and b == 'sp':
    return 'lmp'
  elif a == 'wp' and b == 'vsp':
    return 'hmp'
  elif a == 'wp' and b == 'ep':
    return 'sp'
  ######################
  if a == 'lmp' and b == 'vwp':
    return 'wp'
  elif a == 'lmp' and b == 'wp':
    return 'wp'
  elif a == 'lmp' and b == 'lmp':
    return 'lmp'
  elif a == 'lmp' and b == 'hmp':
    return 'lmp'
  elif a == 'lmp' and b == 'sp':
    return 'hmp'
  elif a == 'lmp' and b == 'vsp':
    return 'hmp'
  elif a == 'lmp' and b == 'ep':
    return 'sp'
  #####################
  if a == 'hmp' and b == 'vwp':
    return 'wp'
  elif a == 'hmp' and b == 'wp':
    return 'lmp'
  elif a == 'hmp' and b == 'lmp':
    return 'hmp'
  elif a == 'hmp' and b == 'hmp':
    return 'hmp'
  elif a == 'hmp' and b == 'sp':
    return 'sp'
  elif a == 'hmp' and b == 'vsp':
    return 'sp'
  elif a == 'hmp' and b == 'ep':
    return 'vsp'
  ########################
  if a == 'sp' and b == 'vwp':
    return 'lmp'
  elif a == 'sp' and b == 'wp':
    return 'lmp'
  elif a == 'sp' and b == 'lmp':
    return 'hmp'
  elif a == 'sp' and b == 'hmp':
    return 'hmp'
  elif a == 'sp' and b == 'sp':
    return 'sp'
  elif a == 'sp' and b == 'vsp':
    return 'sp'
  elif a == 'sp' and b == 'ep':
    return 'vsp'
  ########################
  if a == 'vsp' and b == 'vwp':
    return 'hmp'
  elif a == 'vsp' and b == 'wp':
    return 'hmp'
  elif a == 'vsp' and b == 'lmp':
    return 'hmp'
  elif a == 'vsp' and b == 'hmp':
    return 'sp'
  elif a == 'vsp' and b == 'sp':
    return 'sp'
  elif a == 'vsp' and b == 'vsp':
    return 'vsp'
  elif a == 'vsp' and b == 'ep':
    return 'vsp'
  #########################
  if a == 'ep' and b == 'vwp':
    return 'hmp'
  elif a == 'ep' and b == 'wp':
    return 'sp'
  elif a == 'ep' and b == 'lmp':
    return 'sp'
  elif a == 'ep' and b == 'hmp':
    return 'sp'
  elif a == 'ep' and b == 'sp':
    return 'vsp'
  elif a == 'ep' and b == 'vsp':
    return 'vsp'
  elif a == 'ep' and b == 'ep':
    return 'ep'

lis = []
for i in range(0,100):
  lis.append(MamdaniInference(i))

def defuzzification(x):
  defuz = {}
  for j in range (0, 100):
    model = 0
    luas = 0
  ################################################
    try :
      if x[j]['fuzOut']['vwp']>0 :
        a = (5*x[j]['fuzOut']['vwp'])+(10*x[j]['fuzOut']['vwp'])+(15*x[j]['fuzOut']['vwp'])
        b = x[j]['fuzOut']['vwp']*3
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    try :
      if x[j]['fuzOut']['wp']>0 :
        a = (15*x[j]['fuzOut']['wp'])+(20*x[j]['fuzOut']['wp'])+(25*x[j]['fuzOut']['wp'])+(30*x[j]['fuzOut']['wp'])+(35*x[j]['fuzOut']['wp'])+(40*x[j]['fuzOut']['wp'])
        b = x[j]['fuzOut']['wp']*6
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    try :
      if x[j]['fuzOut']['lmp']>0 :
        a = (30*x[j]['fuzOut']['lmp'])+(35*x[j]['fuzOut']['lmp'])+(40*x[j]['fuzOut']['lmp'])+(45*x[j]['fuzOut']['lmp'])
        b = x[j]['fuzOut']['lmp']*5
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    try :
      if x[j]['fuzOut']['hmp']>0 :
        a = (40*x[j]['fuzOut']['hmp'])+(45*x[j]['fuzOut']['hmp'])+(50*x[j]['fuzOut']['hmp'])+(55*x[j]['fuzOut']['hmp'])+(60*x[j]['fuzOut']['hmp'])+(65*x[j]['fuzOut']['hmp'])+(70*x[j]['fuzOut']['hmp'])
        b = x[j]['fuzOut']['hmp']*7
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    try :
      if x[j]['fuzOut']['sp']>0 :
        a = (65*x[j]['fuzOut']['sp'])+(70*x[j]['fuzOut']['sp'])+(75*x[j]['fuzOut']['sp'])
        b = x[j]['fuzOut']['sp']*3
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    try :
      if x[j]['fuzOut']['vsp']>0 :
        a = (70*x[j]['fuzOut']['vsp'])+(75*x[j]['fuzOut']['vsp'])+(80*x[j]['fuzOut']['vsp'])+(85*x[j]['fuzOut']['vsp'])+(90*x[j]['fuzOut']['vsp'])
        b = x[j]['fuzOut']['vsp']*5
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    try :
      if x[j]['fuzOut']['ep']>0 :
        a = (80*x[j]['fuzOut']['ep'])+(85*x[j]['fuzOut']['ep'])+(90*x[j]['fuzOut']['ep'])+(95*x[j]['fuzOut']['ep'])+(100*x[j]['fuzOut']['ep'])
        b = x[j]['fuzOut']['ep']*5
        model = model + a
        luas = luas + b
    except:
      pass
    ################################################
    hasil = model/luas
    defuz[lis[j]['id']]=hasil
  return defuz

#disimpan sementara untuk di sorting
temp=defuzzification(lis)

#Sorting hasil dari defuzzification
hasil_sort=sorted(temp.values(),reverse=True )
sorted_dict = {}
for i in hasil_sort:
    for k in temp.keys():
        if temp[k] == i:
            sorted_dict[k] = temp[k]

#Ambil 10 terbaik
def take(n, i):
    return list(islice(i, n))

final = take(10, sorted_dict.items())

#export to excel
wb = Workbook()
sheet1 = wb.add_sheet('Sheet 1')

for i in range(0,10):
  sheet1.write(i, 0, final[i][0])

wb.save('peringkat.xls')